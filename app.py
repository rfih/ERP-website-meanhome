from flask import Flask, render_template, request, jsonify, session, send_file
from datetime import datetime, timezone, date, timedelta
from db import Base, engine, SessionLocal
from models import Order, Task, TaskHistory
from io import BytesIO
import openpyxl
from openpyxl import Workbook
import warnings
from sqlalchemy import and_, text, or_, func, desc, asc, case
try:
    from models import Order, SubTask as Task, StationSchedule
except ImportError:
    from models import Order, Task as Task, StationSchedule
import re, unicodedata
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    from zoneinfo import ZoneInfo                # Python 3.9+
    LOCAL_TZ = ZoneInfo("Asia/Taipei")           # <- change if needed
except Exception:
    LOCAL_TZ = timezone(timedelta(hours=8))      # fallback if tzdata missing

def to_local(dt):
    """Convert any DB timestamp (str or dt) to local tz."""
    udt = as_utc(dt)
    return udt.astimezone(LOCAL_TZ) if udt else None

def ts_local(dt):
    """Format a DB timestamp for JSON/UI in local time."""
    ldt = to_local(dt)
    return ldt.strftime("%Y-%m-%d %H:%M:%S") if ldt else None

def utcnow():
    return datetime.now(timezone.utc)

task_timers = {}

app = Flask(__name__)
app.config.from_object("config")

# Create tables on first run (we'll add Alembic later)
Base.metadata.create_all(bind=engine)

def as_utc(dt):
    """Return a timezone-aware UTC datetime from DB value (datetime or str)."""
    if not dt:
        return None
    if isinstance(dt, str):
        s = dt.replace('Z', '+00:00')           # accept Z suffix
        try:
            dt = datetime.fromisoformat(s)
        except Exception:
            # last-resort parse; treat as UTC
            try:
                dt = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S.%f")
            except Exception:
                dt = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")
            dt = dt.replace(tzinfo=timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt

def recalc_active_groups(order):
    # derive from tasks where remaining > 0
    groups = []
    for t in order.tasks:
        remaining = max(0, (t.quantity or 0) - (t.completed or 0))
        if remaining > 0 and t.group and t.group not in groups:
            groups.append(t.group)
    return groups

def order_progress(order):
    total_q = sum((t.quantity or 0) for t in order.tasks)
    done_q = sum((t.completed or 0) for t in order.tasks)
    return (done_q, total_q)

from datetime import date
from sqlalchemy.sql import func

@app.route("/")
def home():
    qtext = (request.args.get("q") or "").strip()
    dfield = (request.args.get("dfield") or "any").strip()
    dfrom  = (request.args.get("dfrom")  or "").strip()
    dto    = (request.args.get("dto")    or "").strip()

    df = _ymd_digits(dfrom)
    dt = _ymd_digits(dto)
    if df and not dt:
        dt = df

    def norm_col(col):
        return func.replace(func.replace(col, '-', ''), '/', '')

    def date_cond_for(col):
        c = norm_col(col)
        preds = []
        if df: preds.append(c >= df)
        if dt: preds.append(c <= dt)
        return and_(*preds) if preds else None

    today_iso = date.today().isoformat()  # <<< use ISO: YYYY-MM-DD

    with SessionLocal() as session:
        q = (session.query(Order)
             .filter(or_(Order.archived == False, Order.archived.is_(None))))

        if qtext:
            kw = f"%{qtext}%"
            q = (q.outerjoin(Task)
                   .filter(or_(
                       Order.manufacture_code.ilike(kw),
                       Order.customer.ilike(kw),
                       Order.product.ilike(kw),
                       Order.jobdesc.ilike(kw),
                       Task.task.ilike(kw),
                       Task.group.ilike(kw)
                   ))
                   .distinct())

        if df or dt:
            if dfield == "demand":
                cond = date_cond_for(Order.demand_date)
                if cond is not None: q = q.filter(cond)
            elif dfield == "delivery":
                cond = date_cond_for(Order.delivery)
                if cond is not None: q = q.filter(cond)
            elif dfield == "datecreate":
                cond = date_cond_for(Order.datecreate)
                if cond is not None: q = q.filter(cond)
            else:
                conds = [c for c in [
                    date_cond_for(Order.demand_date),
                    date_cond_for(Order.delivery),
                    date_cond_for(Order.datecreate),
                ] if c is not None]
                if conds:
                    q = q.filter(or_(*conds))

        orders = q.order_by(Order.id.desc()).all()

        vm = []
        for o in orders:
            done_q, total_q = order_progress(o)

            sub_rows = []
            for t in o.tasks:
                # normalize StationSchedule.plan_date to DATE in SQL, compare with today
                planned_today = session.query(StationSchedule.id).filter(
                    StationSchedule.task_id == t.id,
                    func.date(
                        func.replace(func.replace(StationSchedule.plan_date, '/', '-'), '.', '-')
                    ) == today_iso
                ).first() is not None

                sub_rows.append({
                    "id": t.id,
                    "group": t.group or "",
                    "task": t.task or "",
                    "quantity": t.quantity or 0,
                    "completed": t.completed or 0,
                    "remaining": max(0, (t.quantity or 0) - (t.completed or 0)),
                    "running": bool(getattr(t, "start_time", None) and not getattr(t, "stop_time", None)),
                    "started_at": (t.start_time.isoformat() if getattr(t, "start_time", None) else None),
                    "planned_today": planned_today,   # <<< NEW
                })

            assigned_today = assigned_groups_for_order_today(session, o.id, today_iso)
            today_set, carry_set = assigned_groups_for_order_rolling(session, o.id, today_iso)
            # Build display: today first, then carryover with label
            assigned_list = []
            for g in sorted(today_set):
                if g: assigned_list.append(g)
            for g in sorted(carry_set - today_set):
                if g: assigned_list.append(f"{g}（延續）")
            assigned_groups_disp = "、".join(assigned_list)
            
            vm.append({
                "id": o.id,
                "manufacture_code": o.manufacture_code or "",
                "customer": o.customer or "",
                "product": o.product or "",
                "demand_date": o.demand_date or "",
                "delivery": o.delivery or "",
                "quantity": o.quantity or 0,
                "datecreate": o.datecreate or "",
                "fengbian": o.fengbian or "",
                "wallpaper": o.wallpaper or "",
                "jobdesc": o.jobdesc or "",
                "assigned_groups": assigned_groups_disp,
                "progress": (done_q, total_q),
                "sub_tasks": sub_rows,
            })

        return render_template(
            "index.html",
            orders=vm,
            today_date=today_iso,
            archived_page=False,
            qtext=qtext
        )
    
@app.route("/archived")
def archived_page():
    qtext = (request.args.get("q") or "").strip()
    dfield = (request.args.get("dfield") or "any").strip()   # any | demand | delivery | datecreate
    dfrom  = (request.args.get("dfrom")  or "").strip()
    dto    = (request.args.get("dto")    or "").strip()

    df = _ymd_digits(dfrom)
    dt = _ymd_digits(dto)
    if df and not dt:
        dt = df  # single-day filter if only start is provided

    def norm_col(col):
        # turn 'YYYY/MM/DD' or 'YYYY-MM-DD' into 'YYYYMMDD' on the SQL side
        return func.replace(func.replace(col, '-', ''), '/', '')

    def date_cond_for(col):
        c = norm_col(col)
        preds = []
        if df: preds.append(c >= df)
        if dt: preds.append(c <= dt)
        return and_(*preds) if preds else None
    today = datetime.today().strftime("%Y/%m/%d")
    with SessionLocal() as session:
        q = session.query(Order).filter(Order.archived == True)

        if qtext:
            kw = f"%{qtext}%"
            q = (q.outerjoin(Task)
                   .filter(or_(
                       Order.manufacture_code.ilike(kw),
                       Order.customer.ilike(kw),
                       Order.product.ilike(kw),
                       Order.jobdesc.ilike(kw),
                       Task.task.ilike(kw),
                       Task.group.ilike(kw)
                   ))
                   .distinct())
            
        # date filter (NEW)
        if df or dt:
            if dfield == "demand":
                cond = date_cond_for(Order.demand_date)
                if cond is not None: q = q.filter(cond)
            elif dfield == "delivery":
                cond = date_cond_for(Order.delivery)
                if cond is not None: q = q.filter(cond)
            elif dfield == "datecreate":
                cond = date_cond_for(Order.datecreate)
                if cond is not None: q = q.filter(cond)
            else:  # any
                conds = [c for c in [
                    date_cond_for(Order.demand_date),
                    date_cond_for(Order.delivery),
                    date_cond_for(Order.datecreate),
                ] if c is not None]
                if conds:
                    q = q.filter(or_(*conds))


        orders = q.order_by(Order.id.desc()).all()

        vm = []
        for o in orders:
            done_q, total_q = order_progress(o)
            vm.append({
                "id": o.id,
                "manufacture_code": o.manufacture_code or "",
                "customer": o.customer or "",
                "product": o.product or "",
                "demand_date": o.demand_date or "",
                "delivery": o.delivery or "",
                "quantity": o.quantity or 0,
                "datecreate": o.datecreate or "",
                "fengbian": o.fengbian or "",
                "wallpaper": o.wallpaper or "",
                "jobdesc": o.jobdesc or "",
                "active_groups": ", ".join(recalc_active_groups(o)),
                "progress": (done_q, total_q),
                "sub_tasks": [
                    {
                        "id": t.id,
                        "group": t.group or "",
                        "task": t.task or "",
                        "quantity": t.quantity or 0,
                        "completed": t.completed or 0,
                        "remaining": max(0, (t.quantity or 0) - (t.completed or 0)),
                        "running": bool(getattr(t, "start_time", None) and not getattr(t, "stop_time", None)),
                        "started_at": (t.start_time.isoformat() if getattr(t, "start_time", None) else None),
                    } for t in o.tasks
                ]
            })
        return render_template("index.html", orders=vm, today_date=today, archived_page=True, qtext=qtext, dfield=dfield, dfrom=dfrom, dto=dto)


@app.route("/add_order", methods=["POST"])
def add_order():
    data = request.get_json()
    with SessionLocal() as session:
        o = Order(
            manufacture_code=data.get("manufacture_code"),
            customer=data.get("customer"),
            product=data.get("product"),
            demand_date=data.get("demand_date"),
            delivery=data.get("delivery"),
            quantity=int(data.get("quantity", 0)),
            datecreate=data.get("datecreate"),
            fengbian=data.get("fengbian"),
            wallpaper=data.get("wallpaper"),
            jobdesc=data.get("jobdesc"),
        )
        session.add(o)
        session.flush()  # get o.id before commit

        # Create tasks from planned stations
        for t in data.get("initial_tasks", []):
            session.add(Task(
                order_id=o.id,
                group=t.get("group"),
                task=t.get("task"),
                quantity=o.quantity,
                completed=0
            ))

        session.commit()
        return jsonify(success=True, order_id=o.id)

@app.route("/add_task", methods=["POST"])
def add_task():
    data = request.get_json()
    with SessionLocal() as session:
        order_id = int(data["order_id"])
        o = session.get(Order, order_id)
        if not o:
            return jsonify(success=False, message="Order not found"), 404

        t = Task(
            order_id=o.id,
            group=data.get("group"),
            task=data.get("task"),
            quantity=int(data.get("quantity", 0)),
            completed=int(data.get("completed", 0)),
        )
        session.add(t)
        session.commit()
        return jsonify(success=True, task_id=t.id)

@app.route("/get_task/<int:task_id>")
def get_task(task_id):
    with SessionLocal() as session:
        t = session.get(Task, task_id)
        if not t:
            return jsonify(success=False, message="Task not found"), 404
        return jsonify(success=True, task={
            "id": t.id, "order_id": t.order_id, "group": t.group or "",
            "task": t.task or "", "quantity": t.quantity or 0, "completed": t.completed or 0
        })

@app.route("/update_task", methods=["POST"])
def update_task():
    data = request.get_json()
    tid = int(data.get("task_id", 0))
    new_completed = int(data.get("completed", 0))

    with SessionLocal() as s:
        t = s.get(Task, tid)
        if not t: return jsonify(success=False, message="Task not found"), 404

        qty  = t.quantity or 0
        prev = t.completed or 0
        t.completed = max(0, min(qty, new_completed))
        delta = t.completed - prev

        s.add(TaskHistory(
            task_id=tid, timestamp=utcnow(), note="update",
            completed=t.completed, delta=delta
        ))
        s.commit()
        return jsonify(success=True, completed=t.completed)
    
def iso(dt):
    if not dt:
        return None
    if isinstance(dt, str):
        # 'YYYY-MM-DD HH:MM:SS(.ffffff)?(+00:00)?' -> ISO + Z
        s = dt.strip().replace(' ', 'T')
        if s.endswith('+00:00'):
            s = s[:-6] + 'Z'
        elif s[-1].isdigit():  # no tz info, assume UTC
            s += 'Z'
        return s
    # datetime
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat().replace("+00:00", "Z")

@app.route("/task-history/<int:task_id>")
def task_history(task_id):
    with SessionLocal() as s:
        rows = (s.query(TaskHistory)
                  .filter(TaskHistory.task_id == task_id)
                  .order_by(TaskHistory.timestamp.desc())
                  .all())

        def canon(note):
            n = (note or "").strip().lower()
            if n in ("變更", "", "update"): return "update"
            if n in ("start", "開始", "開始計時"): return "start"
            if n in ("stop", "停止", "停止計時"): return "stop"
            if n in ("finish", "完成", "結束"):   return "finish"
            if n in ("remove","unplan"):         return "remove"
            if n in ("plan",):                   return "plan"
            return n or "update"

        # Precompute minutes for each STOP by scanning in chronological order
        rows_asc = sorted(rows, key=lambda r: as_utc(r.timestamp) or datetime.min.replace(tzinfo=timezone.utc))
        minutes_map = {}
        last_start_ts = None
        for h in rows_asc:
            note = canon(getattr(h, "note", None))
            ts = as_utc(getattr(h, "timestamp", None))
            if note == "start":
                last_start_ts = ts
            elif note == "stop":
                if last_start_ts and ts:
                    minutes_map[h.id] = int((ts - last_start_ts).total_seconds() // 60)
                last_start_ts = None

        # Build response (still newest-first), formatting timestamps to local time
        out = []
        for h in rows:
            out.append({
                "timestamp": ts_local(h.timestamp),           # use helpers you added
                "note": canon(getattr(h, "note", None)),
                "delta": getattr(h, "delta", None),
                "minutes": minutes_map.get(h.id),             # computed if it's a STOP
                "user": getattr(h, "user", None),
            })
        return jsonify(out)
    
@app.route("/delete_task", methods=["POST"])
def delete_task():
    data = request.get_json()
    order_id = int(data["order_id"])
    task_id = int(data["task_id"])
    with SessionLocal() as session:
        t = session.query(Task).filter_by(id=task_id, order_id=order_id).first()
        if not t:
            return jsonify(success=False, message="Task not found"), 404
        session.delete(t)
        session.commit()
        return jsonify(success=True)

@app.route("/delete_order", methods=["POST"])
def delete_order():
    data = request.get_json()
    order_id = int(data["order_id"])
    with SessionLocal() as session:
        o = session.get(Order, order_id)
        if not o:
            return jsonify(success=False, message="Order not found"), 404
        session.delete(o)
        session.commit()
        return jsonify(success=True)

@app.route("/update_order", methods=["POST"])
def update_order():
    data = request.get_json()
    oid = int(data.get("order_id", 0))
    if not oid:
        return jsonify(success=False, message="order_id 缺少"), 400

    with SessionLocal() as session:
        o = session.get(Order, oid)
        if not o:
            return jsonify(success=False, message="Order not found"), 404

        # update fields
        o.manufacture_code = data.get("manufacture_code") or ""
        o.customer = data.get("customer") or ""
        o.product = data.get("product") or ""
        o.demand_date = data.get("demand_date") or ""
        o.delivery = data.get("delivery") or ""
        try:
            o.quantity = int(data.get("quantity") or 0)
        except Exception:
            o.quantity = 0
        o.datecreate = data.get("datecreate") or ""
        o.fengbian = data.get("fengbian") or ""
        o.wallpaper = data.get("wallpaper") or ""
        o.jobdesc = data.get("jobdesc") or ""

        session.commit()

        # send updated object back
        return jsonify(success=True, order={
            "id": o.id,
            "manufacture_code": o.manufacture_code,
            "customer": o.customer,
            "product": o.product,
            "demand_date": o.demand_date,
            "delivery": o.delivery,
            "quantity": o.quantity,
            "datecreate": o.datecreate,
            "fengbian": o.fengbian,
            "wallpaper": o.wallpaper,
            "jobdesc": o.jobdesc,
        })
    
@app.route("/update_task_info", methods=["POST"])
def update_task_info():
    data = request.get_json()
    task_id = int(data["task_id"])
    order_id = int(data["order_id"])

    with SessionLocal() as session:
        t = session.query(Task).filter_by(id=task_id, order_id=order_id).first()
        if not t:
            return jsonify(success=False, message="Task not found"), 404
        t.group = data.get("group", t.group)
        t.task = data.get("task", t.task)
        t.quantity = int(data.get("quantity", t.quantity or 0))
        session.commit()
        return jsonify(success=True)
    
@app.route("/update_task_timer", methods=["POST"])
def update_task_timer():
    data = request.get_json()
    tid = int(data.get("task_id", 0))
    action = (data.get("action") or "").lower()

    with SessionLocal() as s:
        t = s.get(Task, tid)
        if not t:
            return jsonify(success=False, message="Task not found"), 404

        now = utcnow()

        if action == "start":
            t.start_time = now
            t.stop_time = None
            s.add(TaskHistory(task_id=tid, timestamp=now, note="start"))
            s.commit()
            return jsonify(success=True, running=True)

        elif action == "stop":
            t.stop_time = now
            # just record a stop row; no minutes column on DB
            s.add(TaskHistory(task_id=tid, timestamp=now, note="stop"))
            # if you still want to show minutes to the UI, compute it here too:
            mins = 0
            if getattr(t, "start_time", None):
                try:
                    mins = int((as_utc(now) - as_utc(t.start_time)).total_seconds() // 60)
                except Exception:
                    mins = 0
            s.commit()
            return jsonify(success=True, running=False, minutes=mins)

        else:
            return jsonify(success=False, message="Invalid action"), 400

with open("schema.sql", "r", encoding="utf-8") as f:
    schema = f.read()

@app.route("/export/orders.xlsx")
def export_orders_xlsx():
    wb = Workbook()
    ws_orders = wb.active
    ws_orders.title = "orders"

    # Header row (match your UI)
    ws_orders.append(["客戶交期","預計出貨","製令號碼","專案名稱","產品名稱","數量","派單日","封邊","面飾","作業內容","order_id"])

    with SessionLocal() as s:
        orders = s.query(Order).order_by(Order.id).all()
        for o in orders:
            ws_orders.append([
                o.demand_date or "",
                o.delivery or "",
                o.manufacture_code or "",
                o.customer or "",
                o.product or "",
                o.quantity or 0,
                o.datecreate or "",
                o.fengbian or "",
                o.wallpaper or "",
                o.jobdesc or "",
                o.id,
            ])

        # tasks sheet
        ws_tasks = wb.create_sheet("tasks")
        ws_tasks.append(["order_id","組別","工作內容","數量","已完成","task_id"])
        tasks = s.query(Task).order_by(Task.order_id, Task.id).all()
        for t in tasks:
            ws_tasks.append([t.order_id, t.group or "", t.task or "", t.quantity or 0, t.completed or 0, t.id])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="orders_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def _cell(v):
    # keep as string; your DB stores text dates already
    return "" if v is None else str(v).replace("\n","").strip()

def _norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s))
    return s.replace("\n","").replace(" ","").strip()

def _find_header(ws):
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [_norm(v) for v in row]
        if any(vals):
            return vals, r_idx
    return None, None

# --- flexible patterns (substring/regex) ---
PATTERNS = {
    # demand date: 客戶要求交期 / 客戶交期
    "demand":   [r"客戶.*交期"],
    # delivery: 任意含「出貨/出货」，covers 預計可出貨日 / 預計出貨 / 可出貨日 ...
    "delivery": [r"出貨", r"出货"],
    # code: 製令(號碼/号码) with or without line breaks
    "code":     [r"製令.*(號碼|号码)"],
    "customer": [r"專案名稱"],
    "product":  [r"產品.*名"],
    "qty":      [r"數量"],
    "datecre":  [r"派單日|派单日"],
    "fengbian": [r"封邊|封边"],
    "wall":     [r"面飾|面饰"],
    "jobdesc":  [r"作業內容|作业.*容"],
}

def _col_by_patterns(header, patterns):
    for i, h in enumerate(header):
        for pat in patterns:
            if re.search(pat, h):
                return i
    return None

@app.route("/import_orders", methods=["POST"])
def import_orders():
    f = request.files.get("file")
    if not f:
        return jsonify(success=False, message="No file"), 400

    # take explicit sheet name from query (?sheet=生產中門框) or form field
    sheet_name = (request.args.get("sheet") or request.form.get("sheet") or "").strip()

    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    # choose sheet without scanning all sheetnames if user provided one
    if sheet_name:
        try:
            ws = wb[sheet_name]
        except KeyError:
            return jsonify(success=False, message=f"找不到工作表: {sheet_name}"), 400
    else:
        # default: first sheet (no sheet scanning)
        ws = wb[wb.sheetnames[0]]

    header, header_row = _find_header(ws)
    if not header:
        return jsonify(success=False, message="找不到標題列"), 400

    # acceptable synonyms
    synonyms = {
        "demand":   {"客戶交期","客戶要求交期"},
        "delivery": {"預計出貨","可出貨日"},
        "code":     {"製令號碼","製令号碼","製令号码"},
        "customer": {"專案名稱"},
        "product":  {"產品名稱","產品名"},
        "qty":      {"數量"},
        "datecre":  {"派單日","派单日"},
        "fengbian": {"封邊","封边"},
        "wall":     {"面飾","面饰"},
        "jobdesc":  {"作業內容","作业內容","作业内容"},
    }

    def col(keys):
        for k in keys:
            if k in header:
                return header.index(k)
        return None

    # make a header index on THIS sheet only
    idx = {k: _col_by_patterns(header, v) for k, v in PATTERNS.items()}

    required_keys = ["demand","delivery","code","customer","product","qty","datecre"]
    missing = [k for k in required_keys if idx[k] is None]
    if missing:
        # friendlier error showing which ones are missing
        zh = {
            "demand":"客戶交期",
            "delivery":"預計出貨",
            "code":"製令號碼",
            "customer":"專案名稱",
            "product":"產品名稱",
            "qty":"數量",
            "datecre":"派單日"
        }
        miss_names = "、".join(zh[k] for k in missing)
        return jsonify(success=False, message=f"此工作表缺少必要欄位：{miss_names}（{sheet_name or ws.title}）"), 400

    def cell(v):
        return "" if v is None else str(v).strip()

    created = 0
    with SessionLocal() as s:
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            if not row or not any(row): 
                continue
            def get(key):
                i = idx[key]
                return cell(row[i]) if i is not None and i < len(row) else ""

            try:
                qty = int((get("qty") or "0"))
            except ValueError:
                qty = 0

            o = Order(
                demand_date      = get("demand"),
                delivery         = get("delivery"),
                manufacture_code = get("code"),
                customer         = get("customer"),
                product          = get("product"),
                quantity         = qty,
                datecreate       = get("datecre"),
                fengbian         = get("fengbian"),
                wallpaper        = get("wall"),
                jobdesc          = get("jobdesc"),
            )
            s.add(o)
            created += 1

        s.commit()

    return jsonify(success=True, created=created, sheet=sheet_name or ws.title)

def ensure_archive_cols():
    with engine.connect() as c:
        cols = [r[1] for r in c.execute(text("PRAGMA table_info(orders)")).fetchall()]
        if "archived" not in cols:
            c.execute(text("ALTER TABLE orders ADD COLUMN archived INTEGER DEFAULT 0"))
        if "archived_at" not in cols:
            c.execute(text("ALTER TABLE orders ADD COLUMN archived_at TEXT"))
ensure_archive_cols()

@app.route("/archive_order", methods=["POST"])
def archive_order():
    data = request.get_json()
    oid = int(data.get("order_id", 0))
    do_archive = bool(data.get("archive", True))
    with SessionLocal() as s:
        o = s.get(Order, oid)
        if not o:
            return jsonify(success=False, message="Order not found"), 404
        o.archived = do_archive
        o.archived_at = datetime.now(timezone.utc) if do_archive else None
        s.commit()
        return jsonify(success=True, archived=o.archived)
    
def _ymd_digits(s: str) -> str:
    """Keep only digits, e.g. '2025-08-20' or '2025/8/2' -> '20250820' (no zero-pad for month/day if user typed that, but ok)."""
    if not s: return ""
    return re.sub(r"\D", "", s)

def ensure_station_schedule_table():
    with engine.connect() as c:
        existing = [r[0] for r in c.execute(text("SELECT name FROM sqlite_master WHERE type='table'")).fetchall()]
        if "station_schedule" not in existing:
            c.execute(text("""
              CREATE TABLE station_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                station TEXT,
                task_id INTEGER,
                order_id INTEGER,
                plan_date TEXT,
                planned_qty INTEGER DEFAULT 0,
                sequence INTEGER DEFAULT 0,
                note TEXT DEFAULT '',
                locked INTEGER DEFAULT 0,
                created_at TEXT
              )
            """))
ensure_station_schedule_table()

@app.route("/stations")
def stations():
    # inputs
    group = (request.args.get("group") or "").strip()
    day   = (request.args.get("date")  or date.today().isoformat()).strip()

    with SessionLocal() as s:
        # list of stations (distinct task.group)
        station_list = [row[0] for row in s.query(Task.group)
                                      .filter(Task.group != None)
                                      .distinct().order_by(Task.group).all()]
        if not group and station_list:
            group = station_list[0]

        # if there is a plan for this station+day, show only that (ordered)
        plan_rows = (s.query(StationSchedule)
                      .filter(StationSchedule.station==group,
                              StationSchedule.plan_date==day)
                      .order_by(asc(StationSchedule.sequence), asc(StationSchedule.id))
                      .all())

        items = []
        if plan_rows:
            # join schedule → (task, order)
            task_map = {t.id:(t,o) for t,o in s.query(Task, Order)
                                               .join(Order, Task.order_id==Order.id)
                                               .filter(Task.id.in_([p.task_id for p in plan_rows]),
                                                       or_(Order.archived==False, Order.archived.is_(None)))
                                               .all()}
            for p in plan_rows:
                t,o = task_map.get(p.task_id, (None,None))
                if not t or not o: continue
                # attach last history quickly if you want (optional)
                items.append({"task": t, "order": o, "plan": p})
        else:
            # no plan yet → show backlog for that station (active orders only)
            q = (s.query(Task, Order)
                   .join(Order, Task.order_id==Order.id)
                   .filter(Task.group==group,
                           or_(Order.archived==False, Order.archived.is_(None)))
                   .order_by(desc(Order.id)))
            items = [{"task":t, "order":o, "plan":None} for t,o in q]

        # today label for header
        today_str = date.today().isoformat().replace("-","/")

        return render_template("stations.html",
                               station_list=station_list,
                               selected_group=group,
                               tasks=items,
                               today_date=today_str)
    
@app.route("/stations/publish", methods=["POST"])
def stations_publish():
    data = request.get_json()
    day = data.get("date")
    station = data.get("station")
    task_ids = data.get("task_ids") or []
    default_qty = int(data.get("planned_qty") or 0)

    with SessionLocal() as s:
        # fetch remaining per task to set a sane default if planned_qty not provided
        rem = {t.id: max(0, (t.quantity or 0) - (t.completed or 0))
               for t in s.query(Task).filter(Task.id.in_(task_ids)).all()}

        # get current max sequence
        max_seq = s.query(StationSchedule.sequence)\
                   .filter(StationSchedule.station==station,
                           StationSchedule.plan_date==day)\
                   .order_by(desc(StationSchedule.sequence)).first()
        seq = (max_seq[0] if max_seq else 0)

        for tid in task_ids:
            seq += 1
            s.add(StationSchedule(
                station=station,
                task_id=tid,
                order_id=s.query(Task.order_id).filter(Task.id==tid).scalar(),
                plan_date=day,
                planned_qty= default_qty or rem.get(tid, 0),
                sequence=seq
            ))
        s.commit()
    return jsonify(success=True)

@app.route("/stations/plan/update", methods=["POST"])
def stations_plan_update():
    data = request.get_json()
    updates = data.get("rows") or []  # [{id, planned_qty, sequence}, ...]
    with SessionLocal() as s:
        for u in updates:
            r = s.query(StationSchedule).get(int(u["id"]))
            if not r: continue
            if "planned_qty" in u: r.planned_qty = int(u["planned_qty"])
            if "sequence" in u:    r.sequence    = int(u["sequence"])
        s.commit()
    return jsonify(success=True)

@app.route("/stations/plan/remove", methods=["POST"])
def stations_plan_remove():
    data = request.get_json()
    rid = int(data.get("id"))
    with SessionLocal() as s:
        s.query(StationSchedule).filter(StationSchedule.id==rid).delete()
        s.commit()
    return jsonify(success=True)

@app.route("/finish_task", methods=["POST"])
def finish_task():
    data = request.get_json()
    tid = int(data.get("task_id", 0))

    with SessionLocal() as s:
        t = s.get(Task, tid)
        if not t: return jsonify(success=False, message="Task not found"), 404

        prev = t.completed or 0
        qty  = t.quantity or 0
        t.completed = qty

        s.add(TaskHistory(
            task_id=tid, timestamp=utcnow(), note="finish",
            completed=t.completed, delta=(t.completed - prev)
        ))

        # optional: unplan it everywhere
        try:
            s.query(StationSchedule).filter(StationSchedule.task_id == tid).delete()
        except Exception:
            pass

        s.commit()
        return jsonify(success=True, completed=t.completed, quantity=qty)

    
TZ = timezone(timedelta(hours=8))

def today_yyyy_mm_dd():
    return datetime.now(TZ).strftime("%Y-%m-%d")

def station_names(session):
    # distinct station list from tasks."group"
    rows = session.query(Task.group).filter(Task.group.isnot(None)).distinct().all()
    return [r[0] for r in rows]

def make_item(session, t):
    o = session.get(Order, t.order_id)
    if not o:
        return None
    # skip finished
    qty = t.quantity or 0
    comp = t.completed or 0
    if qty - comp <= 0:
        return None
    return {"order": o, "task": t}

@app.route("/stations/today")
def stations_today():
    group   = (request.args.get("group") or "").strip()
    day     = (request.args.get("date") or today_yyyy_mm_dd()).strip()
    rolling = (request.args.get("rolling") or "1") != "0"  # default rolling
    sort    = (request.args.get("sort") or "seq").strip()  # "seq" | "due"

    with SessionLocal() as s:
        stations = station_names(s)
        if not group and stations:
            group = stations[0]

        tasks = []
        if group:
            base = (
                s.query(Task, StationSchedule.plan_date, StationSchedule.sequence)
                 .join(Order, Order.id == Task.order_id)
                 .join(StationSchedule, StationSchedule.task_id == Task.id)
                 .filter(StationSchedule.station == group)
                 .filter((Task.quantity.isnot(None)) & (Task.quantity > (Task.completed or 0)))
                 .filter(or_(Order.archived == None, Order.archived == 0))
            )
            if not rolling:
                base = base.filter(StationSchedule.plan_date == day)

            if sort == "due":
                base = base.order_by(*effective_due_orderby_asc(),
                                    asc(StationSchedule.plan_date),
                                    asc(StationSchedule.sequence),
                                    asc(Task.id))
            else:
                base = base.order_by(asc(StationSchedule.plan_date),
                                    asc(StationSchedule.sequence),
                                    asc(Task.id))

            rows = base.all()
            for t, plan_date, seq in rows:
                if remaining_of(t) <= 0:
                    continue
                o = s.get(Order, t.order_id)
                # effective due + source label
                eff_date, eff_src = None, None
                if o.delivery and o.demand_date:
                    eff_date = min(o.delivery, o.demand_date)
                    eff_src = "預計出貨" if o.delivery <= o.demand_date else "客戶交期"
                elif o.delivery:
                    eff_date, eff_src = o.delivery, "預計出貨"
                elif o.demand_date:
                    eff_date, eff_src = o.demand_date, "客戶交期"

                lu = last_order_update(s, o.id) if 'last_order_update' in globals() else None

                tasks.append({
                    "order": o,
                    "task": t,
                    "plan_date": plan_date,
                    "sequence": seq,
                    "effective_due": eff_date,
                    "effective_due_src": eff_src,
                    "last_update": lu,
                })

        return render_template(
            "stations.html",
            page_mode="today",
            day=day,
            tasks=tasks,
            station_list=stations,
            selected_group=group,
            today_date=today_yyyy_mm_dd(),
            rolling=rolling,
            sort=sort,
        )


@app.route("/stations/backlog")
def stations_backlog():
    group = (request.args.get("group") or "").strip()
    sort  = (request.args.get("sort") or "due").strip()   # "due" | "fifo"

    with SessionLocal() as s:
        stations = station_names(s)
        if not group and stations:
            group = stations[0]

        tasks = []
        if group:
            eff = effective_due_expr()
            q = (
                s.query(Task)
                 .join(Order, Order.id == Task.order_id)
                 .filter(Task.group == group)
                 .filter((Task.quantity.isnot(None)) & (Task.quantity > (Task.completed or 0)))
                 .filter(or_(Order.archived == None, Order.archived == 0))
            )

            if sort == "due":
                # Use julianday so past dates (earlier) come first; push NULLs last
                q = q.order_by(*effective_due_orderby_asc(), Order.id.asc())
            else:
                q = q.order_by(Order.id.asc())

            for t in q.all():
                if remaining_of(t) <= 0:
                    continue
                o = s.get(Order, t.order_id)
                # effective due + source label
                eff_date, eff_src = None, None
                if o.delivery and o.demand_date:
                    eff_date = min(o.delivery, o.demand_date)
                    eff_src = "預計出貨" if o.delivery <= o.demand_date else "客戶交期"
                elif o.delivery:
                    eff_date, eff_src = o.delivery, "預計出貨"
                elif o.demand_date:
                    eff_date, eff_src = o.demand_date, "客戶交期"

                # latest update across the order (optional; keep if you already have)
                lu = last_order_update(s, o.id) if 'last_order_update' in globals() else None

                tasks.append({
                    "order": o,
                    "task": t,
                    "effective_due": eff_date,
                    "effective_due_src": eff_src,
                    "last_update": lu,
                })

        return render_template(
            "stations.html",
            page_mode="backlog",
            tasks=tasks,
            station_list=stations,
            selected_group=group,
            today_date=today_yyyy_mm_dd(),
            sort=sort,
        )

@app.route("/stations/plan", methods=["POST"])
def stations_plan():
    data = request.get_json() or {}
    station = (data.get("station") or "").strip()
    day = (data.get("date") or today_yyyy_mm_dd()).strip()
    ids = data.get("task_ids") or []

    if not station or not ids:
        return jsonify(success=False, message="Missing station or task_ids"), 400

    added, skipped = [], []
    with SessionLocal() as s:
        max_seq = s.query(func.max(StationSchedule.sequence))\
                   .filter(StationSchedule.station==station,
                           StationSchedule.plan_date==day)\
                   .scalar() or 0

        for tid in ids:
            try:
                tid = int(tid)
            except Exception:
                skipped.append({"id": tid, "reason": "bad_id"})
                continue

            t = s.get(Task, tid)
            if not t:
                skipped.append({"id": tid, "reason": "not_found"})
                continue

            # skip finished
            qty = t.quantity or 0
            comp = t.completed or 0
            if qty - comp <= 0:
                skipped.append({"id": tid, "reason": "finished"})
                continue

            # skip if already planned for that date+station
            exists = s.query(StationSchedule).filter(
                StationSchedule.station==station,
                StationSchedule.plan_date==day,
                StationSchedule.task_id==tid
            ).first()
            if exists:
                skipped.append({"id": tid, "reason": "exists"})
                continue

            max_seq += 10
            s.add(StationSchedule(
                station=station,
                task_id=tid,
                order_id=t.order_id,
                plan_date=day,
                planned_qty=qty - comp,
                sequence=max_seq,
                note=""
            ))
            added.append(tid)

        s.commit()

    return jsonify(success=True, station=station, date=day, added=added, skipped=skipped)

@app.route("/stations/unplan", methods=["POST"])
def stations_unplan():
    data = request.get_json() or {}
    station = (data.get("station") or "").strip()
    day = (data.get("date") or "").strip()
    ids = data.get("task_ids") or []
    if not station or not day or not ids:
        return jsonify(success=False, message="Missing station/date/task_ids"), 400

    removed, missing = [], []
    with SessionLocal() as s:
        for tid in ids:
            try:
                tid = int(tid)
            except Exception:
                missing.append(tid); continue
            q = (s.query(StationSchedule)
                   .filter(StationSchedule.station == station,
                           StationSchedule.plan_date == day,
                           StationSchedule.task_id == tid))
            if q.first():
                q.delete(synchronize_session=False)
                removed.append(tid)
            else:
                missing.append(tid)
        s.commit()
    return jsonify(success=True, station=station, date=day,
                   removed=removed, missing=missing)

def last_order_update(session, order_id):
    """
    Latest TaskHistory across all tasks of an order.
    Returns dict: {from_group, timestamp(ISO), delta or None}
    """
    row = (
        session.query(TaskHistory, Task.group)
        .join(Task, TaskHistory.task_id == Task.id)
        .filter(Task.order_id == order_id)
        .order_by(TaskHistory.timestamp.desc())
        .first()
    )
    if not row:
        return None
    h, from_group = row
    # TaskHistory may or may not have 'delta' set (e.g., stop events)
    delta = getattr(h, "delta", None)
    return {
        "from_group": from_group or "",
        "timestamp": iso(h.timestamp),
        "delta": delta,
    }

def remaining_of(task_obj):
    """Convenience: remaining quantity for a task."""
    q = task_obj.quantity or 0
    c = task_obj.completed or 0
    return max(0, q - c)

def effective_due_expr():
    """
    DB-level expression: earliest non-empty of delivery / demand_date.
    Normalizes '' -> NULL and casts to DATE so comparisons & sorting work,
    including past dates.
    """
    dlv = func.date(func.nullif(Order.delivery, ""))       # TEXT '' -> NULL, then DATE
    dem = func.date(func.nullif(Order.demand_date, ""))    # same for demand_date

    return case(
        (and_(dlv.isnot(None), dem.isnot(None), dlv <= dem), dlv),
        (and_(dlv.isnot(None), dem.isnot(None), dem < dlv), dem),
        (dlv.isnot(None), dlv),
        (dem.isnot(None), dem),
        else_=None
    )

def effective_due_orderby_asc():
    eff = effective_due_expr()
    # SQLite trick: ORDER BY (eff IS NULL) puts NULLs last (0 for non-null, 1 for null)
    return asc(case((eff.is_(None), 1), else_=0)), asc(func.julianday(eff))

def fmtdate(val):
    """Return YYYY-MM-DD for date-like values; empty string if missing."""
    if not val:
        return ""
    # Native date/datetime
    if isinstance(val, date) and not isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime):
        return val.date().strftime("%Y-%m-%d")
    # String inputs
    s = str(val).strip()
    if not s:
        return ""
    # common formats
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    # generic fallback: first 10 chars if it looks like an ISO-like date
    return s[:10]

# Register as Jinja filter
app.jinja_env.filters["fmtdate"] = fmtdate

def assigned_groups_for_order_today(session, order_id, day_iso=None):
    """Return a list of station names that this order is planned for on day_iso."""
    day_iso = day_iso or date.today().isoformat()
    rows = (
        session.query(StationSchedule.station)
        .join(Task, StationSchedule.task_id == Task.id)
        .filter(Task.order_id == order_id)
        # normalize any stored formats to ISO date for comparison
        .filter(func.date(func.replace(func.replace(StationSchedule.plan_date, '/', '-'), '.', '-')) == day_iso)
        .distinct()
        .all()
    )
    return [r[0] for r in rows]

def _iso_day(val):
    """Return 'YYYY-MM-DD' string or ''."""
    if not val:
        return ""
    if isinstance(val, date) and not isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip().replace("/", "-").replace(".", "-")
    return s[:10] if len(s) >= 10 else ""

def assigned_groups_for_order_rolling(session, order_id, today_iso=None):
    """
    Return (today_set, carryover_set) of station names for this order:
      - today_set: planned on today
      - carryover_set: planned on a past day and still unfinished
    """
    today_iso = today_iso or date.today().isoformat()
    rows = (
        session.query(StationSchedule.station,
                      StationSchedule.plan_date,
                      Task.quantity, Task.completed)
        .join(Task, StationSchedule.task_id == Task.id)
        .filter(Task.order_id == order_id)
        .all()
    )

    today_set, carry_set = set(), set()
    for station, plan_date, qty, comp in rows:
        qty = qty or 0
        comp = comp or 0
        if qty <= comp:  # finished -> not active
            continue
        day_iso = _iso_day(plan_date)
        if not day_iso:
            continue
        if day_iso == today_iso:
            today_set.add(station or "")
        elif day_iso < today_iso:
            carry_set.add(station or "")
        # (future-dated plans are ignored for this column)

    return today_set, carry_set

if __name__ == "__main__":
    app.run(debug=True)
